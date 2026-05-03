#!/usr/bin/env python3
"""
Convert a "Finances Sheet.xlsx" workbook into two CSVs that the finance dashboard
can import directly: transactions and budgets.

Usage:
    python scripts/finance_import_xlsx.py "Finances Sheet.xlsx"

Outputs (next to the input file unless --out is given):
    transactions.import.csv
    budgets.import.csv

Then in the dashboard:
  - Transactions tab → Pick CSV → Import (replace) → transactions.import.csv
  - Budget tab → Import CSV → budgets.import.csv

Note: never commit either of these CSVs. They contain personal financial data.
.gitignore already excludes *.private.csv and the source .xlsx.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import os
import sys
import uuid
from pathlib import Path

try:
    import openpyxl  # type: ignore
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# Account values seen in the user's sheet → canonical account name in the dashboard.
ACCOUNT_REMAP = {
    "amex": "Amex",
    "debt": "Debit",   # the sheet uses "Debt" (typo for Debit); remap on import
    "debit": "Debit",
    "cash": "Cash",
    "other": "Other",
}

# Category remaps: source-sheet name → canonical key in the dashboard taxonomy.
# Entries here exist because the spreadsheet uses a slightly different label
# than the dashboard's categories.ts. Add a row whenever you rename in one
# place but not the other.
CATEGORY_REMAP = {
    "ChatGPT": "AI Subscription",       # historical name in the budget tab
    "OneDrive": "One Drive",
    "Car Maintenence": "Car Maintenance",  # spelling fix; dashboard taxonomy uses correct spelling
}

# Categories from the dashboard taxonomy. Kept in sync manually with
# src/components/dashboards/finance/categories.ts. If a transaction's category
# (post-remap) isn't in this set, it's emitted as-is — the dashboard will
# render it under "Uncategorized" until you fix it from the UI.
KNOWN_CATEGORIES = {
    "Groceries", "Eating Out", "Laundry", "Hair Cut", "Rent",
    "Car Payment", "Car Insurance", "Car Maintenance", "Gas",
    "Friends Fun", "Friends Food & Drink", "Dates Fun", "Dates Food & Drink",
    "Personal Fun", "Gifts",
    "Spotify", "AI Subscription", "iCloud", "One Drive", "Therapy",
    "Amazon Subscription", "Railway", "Hulu", "Singing Lessons",
    "College", "Phone Bill", "Health Insurance",
    "Personal Care & Grooming", "Tech & Electronics", "Transportation & Gear",
    "Home & Environment", "Clothing & Accessories", "Books", "Health & Wellness",
    "Other", "Travel", "Charity", "Taxes", "Savings",
    "Car",
}


def iso(d: dt.datetime | dt.date | None) -> str:
    if d is None:
        return ""
    if isinstance(d, dt.datetime):
        d = d.date()
    return d.isoformat()


def parse_transactions(ws) -> list[dict]:
    """
    Spending Log column layout (1-indexed, no reliable header row):
        col 1: Date
        col 2: Item
        col 3: Price
        col 4: Account
        col 5: Category
    """
    out: list[dict] = []
    skipped = 0
    for r in range(2, ws.max_row + 1):
        date = ws.cell(r, 1).value
        item = ws.cell(r, 2).value
        price = ws.cell(r, 3).value
        account = ws.cell(r, 4).value
        category = ws.cell(r, 5).value
        if not item or price is None:
            continue
        if not isinstance(price, (int, float)):
            try:
                price = float(price)
            except (TypeError, ValueError):
                skipped += 1
                continue
        if price <= 0:
            skipped += 1
            continue
        acct_raw = (str(account).strip() if account else "")
        acct = ACCOUNT_REMAP.get(acct_raw.lower(), "Other")
        cat_raw = (str(category).strip() if category else "")
        cat = CATEGORY_REMAP.get(cat_raw, cat_raw)
        out.append({
            "id": str(uuid.uuid4()),
            "date": iso(date),
            "item": str(item).strip(),
            "amount": f"{float(price):.2f}",
            "account": acct,
            "category": cat,
            "notes": "",
            "created_at": dt.datetime.now(dt.timezone.utc).isoformat(),
            "updated_at": dt.datetime.now(dt.timezone.utc).isoformat(),
        })
    if skipped:
        print(f"  skipped {skipped} row(s) with non-numeric or non-positive amounts")
    return out


def parse_budgets(ws) -> list[dict]:
    """
    Current Budget tab, "Budget List View" column (col 1 = name, col 2 = amount).
    Stops at the first contiguous run of non-empty rows after row 3.
    """
    today = dt.date.today().isoformat()
    rows: list[dict] = []
    for r in range(3, ws.max_row + 1):
        name = ws.cell(r, 1).value
        amount = ws.cell(r, 2).value
        if not name or amount is None:
            continue
        if str(name).strip().lower() in {"total", "subtotal"}:
            continue
        try:
            amt = float(amount)
        except (TypeError, ValueError):
            continue
        cat = CATEGORY_REMAP.get(str(name).strip(), str(name).strip())
        rows.append({
            "category": cat,
            "monthly_amount": f"{amt:.2f}",
            "effective_from": today,
        })
    return rows


def main() -> int:
    p = argparse.ArgumentParser(description="Convert Finances Sheet.xlsx into dashboard-importable CSVs.")
    p.add_argument("xlsx", help="Path to the .xlsx file (e.g. 'Finances Sheet.xlsx')")
    p.add_argument("--out", default=None, help="Output directory (defaults to same dir as input)")
    args = p.parse_args()

    src = Path(args.xlsx)
    if not src.exists():
        print(f"ERROR: file not found: {src}", file=sys.stderr)
        return 1

    out_dir = Path(args.out) if args.out else src.parent
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(src, data_only=True)

    if "Spending Log" not in wb.sheetnames:
        print("ERROR: workbook is missing a 'Spending Log' sheet.", file=sys.stderr)
        return 1
    if "Current Budget" not in wb.sheetnames:
        print("WARNING: workbook is missing a 'Current Budget' sheet — skipping budgets.")

    txs = parse_transactions(wb["Spending Log"])
    budgets = parse_budgets(wb["Current Budget"]) if "Current Budget" in wb.sheetnames else []

    # Surface unknown categories so the user can fix taxonomy or remap rules.
    unknown = sorted({t["category"] for t in txs if t["category"] and t["category"] not in KNOWN_CATEGORIES})
    if unknown:
        print(f"  {len(unknown)} unknown category value(s) (will show as 'Uncategorized' in dashboard until fixed):")
        for u in unknown:
            print(f"    - {u!r}")

    tx_path = out_dir / "transactions.import.csv"
    with tx_path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["id","date","item","amount","account","category","notes","created_at","updated_at"])
        w.writeheader()
        w.writerows(txs)
    print(f"  wrote {len(txs)} transactions → {tx_path}")

    if budgets:
        bg_path = out_dir / "budgets.import.csv"
        with bg_path.open("w", encoding="utf-8", newline="") as f:
            w = csv.DictWriter(f, fieldnames=["category","monthly_amount","effective_from"])
            w.writeheader()
            w.writerows(budgets)
        print(f"  wrote {len(budgets)} budget rows → {bg_path}")

    print("\nNext: open /dashboards/finance, then:")
    print("  - Transactions tab → Pick CSV → choose transactions.import.csv → Import (replace)")
    print("  - Budget tab → Import CSV → choose budgets.import.csv")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
